# -*- coding: utf-8 -*-
from __future__ import annotations
import os

from io import BytesIO
from django.http import HttpResponse
from django.conf import settings
from django.contrib.staticfiles import finders
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import stringWidth
from io import BytesIO
from reportlab.pdfgen import canvas
from pypdf import PdfReader, PdfWriter
from django.template.loader import get_template
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import mm  # pratique pour penser en millimètres
from pypdf import PdfReader, PdfWriter
from django.contrib import messages
from openpyxl import Workbook
from django.db.models import Q
from django.http import FileResponse, Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template
from django.utils import timezone
from xhtml2pdf import pisa
from django.utils.html import escape
from agents.models import Agent

from .forms import (
    CommissionForm,
    VivierCreateForm,
    VivierUpdateForm,
    VivierValidationForm,
)
from .models import Commission, PieceJointe, Vivier
from .utils import (
    compute_eligibles_agence,
    extract_direction_from_arbo,
    extract_directions_from_agents,
)

# Police lisible (optionnel si tu as le TTF)
try:
    pdfmetrics.registerFont(
        TTFont(
            "DejaVu",
            os.path.join(settings.BASE_DIR, "static", "fonts", "DejaVuSans.ttf"),
        )
    )
    BASE_FONT = "DejaVu"
except Exception:
    BASE_FONT = "Helvetica"


from reportlab.lib.units import mm  # pratique pour penser en millimètres


def draw_grid(c, W, H, step=10 * mm):
    """
    Dessine une grille + règles X/Y et graduations (tous les 'step').
    Origine (0,0) = bas-gauche. A4 ≈ 210x297mm => 595x842pt.
    """
    c.setStrokeGray(0.85)
    c.setFillGray(0.5)
    c.setLineWidth(0.2)

    # Lignes verticales + graduation en bas
    x = 0
    while x <= W:
        c.line(x, 0, x, H)
        c.setFont(BASE_FONT, 6)
        c.drawString(x + 1, 2, f"{int(x)}")
        x += step

    # Lignes horizontales + graduation à gauche
    y = 0
    while y <= H:
        c.line(0, y, W, y)
        c.setFont(BASE_FONT, 6)
        c.drawString(2, y + 1, f"{int(y)}")
        y += step

    # Axes plus marqués
    c.setStrokeGray(0.3)
    c.setLineWidth(0.6)
    c.line(0, 0, W, 0)  # axe X
    c.line(0, 0, 0, H)  # axe Y


def _render_html_to_pdf(template_name: str, context: dict, filename: str):
    tpl = get_template(template_name)
    html = tpl.render(context)
    buf = BytesIO()
    result = pisa.CreatePDF(src=html, dest=buf, encoding="utf-8")
    if result.err:
        return HttpResponse(html)  # debug: affiche le HTML si erreur CSS
    buf.seek(0)
    resp = HttpResponse(buf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp


def draw_wrapped(c, text, x, y, max_width, leading, font=BASE_FONT, size=10):
    """Dessine le texte sur plusieurs lignes (respecte max_width). Retourne le y final."""
    if not text:
        return y
    words = str(text).split()
    line, lines = "", []
    for w in words:
        test = (line + " " + w).strip()
        if stringWidth(test, font, size) <= max_width:
            line = test
        else:
            if line:
                lines.append(line)
            line = w
    if line:
        lines.append(line)
    c.setFont(font, size)
    cur_y = y
    for ln in lines:
        c.drawString(x, cur_y, ln)
        cur_y -= leading
    return cur_y


def pv_modele_static(request, pk):
    return _render_html_to_pdf("vivier/pv_modele_static.html", {}, "PV_MODELE.pdf")


def _render_html_to_pdf(template_name: str, context: dict, filename: str):
    """
    Rend un template HTML en PDF et le renvoie en HttpResponse inline.
    - IMPORTANT: xhtml2pdf supporte un sous-ensemble de CSS. Privilégier le CSS simple inline.
    """
    template = get_template(template_name)
    html = template.render(context)

    pdf_io = BytesIO()
    # NOTE: avoid external HTTP assets; inline CSS or absolute file paths only
    pisa_status = pisa.CreatePDF(src=html, dest=pdf_io, encoding="utf-8")

    if pisa_status.err:
        # En debug, tu peux retourner le HTML pour voir l'erreur:
        return HttpResponse(html)
    pdf_io.seek(0)
    resp = HttpResponse(pdf_io, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp


def draw_grid(c, W, H, step=20 * mm):
    """Grille de debug (pour caler les coordonnées)."""
    c.setStrokeGray(0.85)
    c.setFillGray(0.5)
    c.setLineWidth(0.2)
    x = 0
    while x <= W:
        c.line(x, 0, x, H)
        c.setFont(BASE_FONT, 6)
        c.drawString(x + 1, 2, f"{int(x)}")
        x += step
    y = 0
    while y <= H:
        c.line(0, y, W, y)
        c.setFont(BASE_FONT, 6)
        c.drawString(2, y + 1, f"{int(y)}")
        y += step
    c.setStrokeGray(0.3)
    c.setLineWidth(0.6)
    c.line(0, 0, W, 0)
    c.line(0, 0, 0, H)


def draw_cross(c, x, y, size=4 * mm):
    """Petit repère en croix pour tester un point exact."""
    c.setStrokeGray(0.2)
    c.setLineWidth(0.8)
    c.line(x - size, y, x + size, y)
    c.line(x, y - size, x, y + size)
    c.setFont(BASE_FONT, 8)
    c.drawString(x + 2, y + 2, f"({int(x)},{int(y)})")


# Police (optionnelle si tu as le TTF ; sinon Helvetica)
try:
    pdfmetrics.registerFont(
        TTFont(
            "DejaVu",
            os.path.join(settings.BASE_DIR, "static", "fonts", "DejaVuSans.ttf"),
        )
    )
    BASE_FONT = "DejaVu"
except Exception:
    BASE_FONT = "Helvetica"


def _merge_overlay_on_pdf(
    template_path: str, draw_fn, out_filename: str = "document.pdf"
):
    # Résoudre via staticfiles → fallback /static/
    abs_template = finders.find(template_path)
    if not abs_template:
        abs_template = os.path.join(
            settings.BASE_DIR, "static", template_path.replace("/", os.sep)
        )
        if not os.path.exists(abs_template):
            raise Http404(f"Template PDF introuvable: {template_path}")

    # Lecture du modèle
    try:
        base_reader = PdfReader(abs_template)
    except Exception as e:
        raise Http404(f"Impossible de lire le PDF modèle: {abs_template} ({e})")
    if not base_reader.pages:
        raise Http404("Le PDF modèle ne contient aucune page.")

    base_page = base_reader.pages[0]
    W = float(base_page.mediabox.width)
    H = float(base_page.mediabox.height)

    # Générer l’overlay
    overlay_buf = BytesIO()
    c = canvas.Canvas(overlay_buf, pagesize=(W, H))
    c.setFont(BASE_FONT, 10)
    draw_fn(c, W, H)  # dessiner aux coordonnées du modèle
    c.save()
    overlay_buf.seek(0)

    # Fusion
    overlay_reader = PdfReader(overlay_buf)
    overlay_page = overlay_reader.pages[0]
    base_page.merge_page(overlay_page)

    out = PdfWriter()
    out.add_page(base_page)
    out_buf = BytesIO()
    out.write(out_buf)
    out_buf.seek(0)

    resp = HttpResponse(out_buf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{out_filename}"'
    return resp


def vivier_pv_non_traj(request, pk):
    v = get_object_or_404(Vivier.objects.using("default"), pk=pk)

    # Récup des données (on garde ta logique)
    rows = []
    eligibles = compute_eligibles_agence(v.FonctionCible) if v.FonctionCible else []
    for r in eligibles:
        ag = r.get("agent")
        if not ag or bool(r.get("trajectoire")):
            continue
        com = (
            Commission.objects.using("default")
            .filter(Vivier=v, Matricule=str(ag.Matricule))
            .first()
        )
        rows.append(
            {
                "matricule": ag.Matricule,
                "nom": ag.Nom,
                "prenom": ag.Prenom,
                "fonction": getattr(ag, "FonctionLibelle", "") or "",
                "affectation": f"{getattr(ag,'AffectationCode','') or ''} — {getattr(ag,'AffectationLibelle','') or ''}",
                "reseau": extract_direction_from_arbo(
                    getattr(ag, "ArborescenceAffectation", "") or ""
                ),
                "sanction": getattr(com, "Sanction", "") or "",
                "pi_n1": getattr(com, "PI_n_1", "") or "",
                "pi_n2": getattr(com, "PI_n_2", "") or "",
                "pi_n3": getattr(com, "PI_n_3", "") or "",
                "avis": getattr(com, "AvisCommission", "") or "",
                "note": getattr(com, "Note", "") or "",
                "decision": getattr(com, "Decision", "") or "",
                "motif": getattr(com, "MotifDecision", "") or "",
                "caractere": getattr(com, "Caractere", "") or "",
            }
        )

    # Fonction de dessin : place les textes aux coordonnées du modèle
    def draw(c, W, H):
        # Exemples de zones (en points) à AJUSTER une seule fois pour coller à ton PDF :
        # NB : Origine (0,0) en bas à gauche. A4 = ~595x842.
        c.setFont(BASE_FONT, 12)
        # En-tête / méta
        c.drawString(50, H - 70, f"Vivier : {v.NumCommission}")
        c.drawString(300, H - 70, f"Fonction cible : {v.FonctionCible}")
        c.drawString(
            50,
            H - 90,
            f"Date : {v.DateCreation.strftime('%d/%m/%Y') if v.DateCreation else ''}",
        )

        # Titres de colonnes (si le modèle a déjà les titres imprimés, ne pas redessiner)
        # c.setFont(BASE_FONT, 10)
        # c.drawString(30, H-130, "Matricule"); c.drawString(90, H-130, "Nom"); ...

        # Lignes du tableau (Trajectoire : Non)
        y = H - 150
        line_h = 16  # interligne
        c.setFont(BASE_FONT, 9)
        for r in rows:
            if (
                y < 60
            ):  # saute à une autre page si besoin (si ton modèle a plusieurs pages, adapter)
                break
            c.drawString(30, y, str(r["matricule"]))
            c.drawString(90, y, r["nom"])
            c.drawString(180, y, r["prenom"])
            c.drawString(260, y, r["fonction"])
            c.drawString(410, y, r["affectation"])
            c.drawString(30, y - 12, f"Réseau: {r['reseau']}")
            c.drawString(260, y - 12, f"Sanction: {r['sanction']}")
            c.drawString(
                410, y - 12, f"PI: {r['pi_n1']} / {r['pi_n2']} / {r['pi_n_3']}"
            )
            c.drawString(30, y - 24, f"Avis: {r['avis']}")
            c.drawString(260, y - 24, f"Décision: {r['decision']}  Note: {r['note']}")
            c.drawString(
                410, y - 24, f"Motif: {r['motif']}  Caractère: {r['caractere']}"
            )
            y -= 40  # saute deux lignes pour aérer

        # Zones de signature si ton modèle les prévoit :
        # c.rect(50, 60, 200, 60)  # exemple de cadre signature

    return _merge_overlay_on_pdf("vivier/pv_template.pdf", draw)


def vivier_pv_debug_grid(request, pk):
    v = get_object_or_404(Vivier.objects.using("default"), pk=pk)

    def draw(c, W, H):
        draw_grid(c, W, H, step=20 * mm)
        c.setFont(BASE_FONT, 12)
        c.drawString(30 * mm, H - 25 * mm, f"Vivier : {v.NumCommission}")

    return _merge_overlay_on_pdf(
        "vivier/fiche_template.pdf", draw, out_filename="DEBUG_grid.pdf"
    )


def vivier_list(request):
    q = (request.GET.get("q") or "").strip()
    qs = (
        Vivier.objects.using("default")
        .all()
        .order_by("-DateCreation", "-NumCommission")
    )
    if q:
        qs = qs.filter(
            Q(NumCommission__icontains=q)
            | Q(FonctionCible__icontains=q)
            | Q(DirectionReseau__icontains=q)
        )
    return render(request, "vivier/vivier_list.html", {"viviers": qs, "q": q})


def _next_num_for_year():
    year = timezone.now().year
    suffix = f"/{year}"
    last = (
        Vivier.objects.using("default")
        .filter(NumCommission__endswith=suffix)
        .values_list("NumCommission", flat=True)
        .order_by("-NumCommission")
        .first()
    )
    n = 0
    if last and "/" in last:
        try:
            n = int(last.split("/")[0])
        except Exception:
            n = 0
    return f"{n+1:03d}/{year}"


def vivier_create(request):
    dir_choices = [("", "— Sélectionner —")] + [
        (d, d) for d in extract_directions_from_agents()
    ]
    initial = {
        "NumCommission": _next_num_for_year(),
        "DateCreation": timezone.now().date(),
    }

    if request.method == "POST":
        form = VivierCreateForm(
            request.POST, direction_choices=dir_choices, initial=initial
        )
        if form.is_valid():
            form.save()  # pas de 'using' sur ModelForm
            messages.success(request, "Vivier créé.")
            return redirect("vivier:list")
        messages.error(request, "Veuillez corriger les erreurs.")
    else:
        form = VivierCreateForm(initial=initial, direction_choices=dir_choices)

    return render(request, "vivier/vivier_form_create.html", {"form": form})


def vivier_update(request, pk):
    """
    Modifier un vivier :
    - Form 'compléments' (Observation + PJ)
    - Form 'validation' (Valide + DateValidation)
    - Tableau des agents éligibles (Oui = autorisés, Non = exceptions)
    """
    v = get_object_or_404(Vivier.objects.using("default"), pk=pk)

    # ---------- FORMS ----------
    if request.method == "POST":
        compl_form = VivierUpdateForm(
            request.POST, request.FILES, instance=v, prefix="c"
        )
        val_form = VivierValidationForm(request.POST, instance=v, prefix="v")

        if compl_form.is_valid() and val_form.is_valid():
            # 1) Observation
            compl_form.save()  # (ModelForm → pas de 'using')

            # 2) Validation (Valide + DateValidation)
            val_form.save()

            messages.success(request, f"Vivier {v.NumCommission} mis à jour.")
            return redirect("vivier:update", pk=v.pk)
        else:
            messages.error(request, "Veuillez corriger les erreurs.")
    else:
        compl_form = VivierUpdateForm(instance=v, prefix="c")
        val_form = VivierValidationForm(instance=v, prefix="v")

    # ---------- AGENTS ÉLIGIBLES ----------
    agents = []
    if v.FonctionCible:
        try:
            # -> liste de dicts: {"agent": Agent, "trajectoire": bool, "exception": bool}
            raw = compute_eligibles_agence(v.FonctionCible)
        except Exception as e:
            messages.warning(
                request,
                f"Impossible de calculer les agents éligibles pour « {v.FonctionCible} » : {e}",
            )
            raw = []

        for r in raw:
            ag = r.get("agent")
            if not ag:
                continue

            # Commission déjà saisie ?
            com = (
                Commission.objects.using("default")
                .filter(Vivier=v, Matricule=str(ag.Matricule))
                .first()
            )

            agents.append(
                {
                    # Identité
                    "agent": ag,
                    "matricule": ag.Matricule,
                    "nom": ag.Nom,
                    "prenom": ag.Prenom,
                    # RH
                    "date_entree": getattr(ag, "DateEntree", None),
                    "anciennete_cpm": getattr(ag, "AncienneteAcquiseCPM", None),
                    # Fonction / affectation
                    "fonction_code": getattr(ag, "FonctionCode", "") or "",
                    "fonction_libelle": getattr(ag, "FonctionLibelle", "") or "",
                    "date_fonction": getattr(ag, "DateEffetFonction", None),
                    "anc_fonction": getattr(ag, "DateEffetFonction", None),
                    "aff_code": getattr(ag, "AffectationCode", "") or "",
                    "aff_lib": getattr(ag, "AffectationLibelle", "") or "",
                    "reseau": extract_direction_from_arbo(
                        getattr(ag, "ArborescenceAffectation", "") or ""
                    ),
                    # Statuts
                    "trajectoire": bool(r.get("trajectoire")),
                    "exception": bool(r.get("exception")),
                    # Commission (si existante)
                    "sanction": getattr(com, "Sanction", "") or "",
                    "pi_n1": getattr(com, "PI_n_1", "") or "",
                    "pi_n2": getattr(com, "PI_n_2", "") or "",
                    "pi_n3": getattr(com, "PI_n_3", "") or "",
                    "avis": getattr(com, "AvisCommission", "") or "",
                    "note": getattr(com, "Note", "") or "",
                    "decision": getattr(com, "Decision", "") or "",
                }
            )

        # Filtre par boutons
        traj = (request.GET.get("traj") or "").lower()
        if traj == "oui":
            agents = [a for a in agents if a["trajectoire"]]
        elif traj == "non":
            agents = [a for a in agents if a["exception"]]
        # sinon : Tous

    return render(
        request,
        "vivier/vivier_form_update.html",
        {
            "v": v,
            "form": compl_form,  # Observation + pièce
            "val_form": val_form,  # Validé + DateValidation
            "agents": agents,
        },
    )


def vivier_delete(request, pk):
    v = get_object_or_404(Vivier.objects.using("default"), pk=pk)
    num = v.NumCommission  # pour le message après suppression
    v.delete(using="default")
    messages.success(request, f"Vivier {num} supprimé.")
    return redirect("vivier:list")


def vivier_piece_delete(request, pk, piece_id):
    v = get_object_or_404(Vivier, pk=pk)
    p = get_object_or_404(PieceJointe, pk=piece_id, vivier=v)  # <-- PieceJointe
    p.delete()
    messages.success(request, "Pièce supprimée.")
    return redirect("vivier:update", pk=v.pk)


def commission_edit(request, vivier_id, matricule):
    v = get_object_or_404(Vivier.objects.using("default"), pk=vivier_id)
    com, _ = Commission.objects.using("default").get_or_create(
        Vivier=v, Matricule=str(matricule)
    )

    # ---- fiche agent (pour affichage dans le template) ----
    ag = Agent.objects.using("default").filter(Matricule=str(matricule)).first()
    agent_ctx = {
        "matricule": str(matricule),
        "nom": getattr(ag, "Nom", "") if ag else "",
        "prenom": getattr(ag, "Prenom", "") if ag else "",
        "fonction": getattr(ag, "FonctionLibelle", "") if ag else "",
        "date_fonction": getattr(ag, "DateEffetFonction", None) if ag else None,
        "affectation_code": getattr(ag, "AffectationCode", "") if ag else "",
        "affectation_libelle": getattr(ag, "AffectationLibelle", "") if ag else "",
    }
    # -------------------------------------------------------

    if request.method == "POST":
        form = CommissionForm(request.POST, instance=com)
        if form.is_valid():
            form.save()
            messages.success(request, "Commission enregistrée.")
            return redirect("vivier:update", pk=v.pk)
        messages.error(request, "Veuillez corriger les erreurs.")
    else:
        form = CommissionForm(instance=com)

    return render(
        request,
        "vivier/commission_form.html",
        {
            "v": v,
            "matricule": matricule,
            "form": form,
            "agent": agent_ctx,  # <-- dispo dans le template
        },
    )


def download_pj(request, pk):
    pj = get_object_or_404(PieceJointe.objects.using("default"), pk=pk)
    if not pj.data:
        raise Http404("Aucune pièce jointe.")
    mime = pj.mime or "application/octet-stream"
    name = pj.nom or f"piece_{pj.pk}"
    return FileResponse(
        BytesIO(pj.data),
        as_attachment=True,
        filename=name,
        content_type=mime,
    )


def _render_pdf_from_template(template_name, context):
    """Rend un template HTML en PDF via xhtml2pdf (simple et Windows-friendly)."""
    html = get_template(template_name).render(context)
    response = HttpResponse(content_type="application/pdf")
    # Nom de fichier proposé au téléchargement
    filename = context.get("pdf_filename", "document.pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    pisa.CreatePDF(src=html, dest=response)  # génère le PDF dans la réponse
    return response


def commission_print(request, vivier_id, matricule):
    v = get_object_or_404(Vivier.objects.using("default"), pk=vivier_id)
    ag = Agent.objects.using("default").filter(Matricule=str(matricule)).first()
    com = (
        Commission.objects.using("default")
        .filter(Vivier=v, Matricule=str(matricule))
        .first()
    )

    agent = {
        "matricule": str(matricule),
        "nom": getattr(ag, "Nom", "") if ag else "",
        "prenom": getattr(ag, "Prenom", "") if ag else "",
        "date_entree": getattr(ag, "DateEntree", None) if ag else None,
        "affectation_code": getattr(ag, "AffectationCode", "") if ag else "",
        "affectation_libelle": getattr(ag, "AffectationLibelle", "") if ag else "",
        "fonction": getattr(ag, "FonctionLibelle", "") if ag else "",
        "date_fonction": getattr(ag, "DateEffetFonction", None) if ag else None,
        "anciennete_cpm": getattr(ag, "AncienneteAcquiseCPM", "") if ag else "",
        "anciennete_fonction": (
            getattr(ag, "AncienneteFonction", "")
            if hasattr(ag, "AncienneteFonction")
            else ""
        ),
    }

    commission = {
        "MoyennePerformance": (
            getattr(com, "MoyennePerformance", "")
            if hasattr(com, "MoyennePerformance")
            else ""
        ),
        "PI_n_1": getattr(com, "PI_n_1", "") if com else "",
        "PI_n_2": getattr(com, "PI_n_2", "") if com else "",
        "PI_n_3": getattr(com, "PI_n_3", "") if com else "",
        "Sanction": getattr(com, "Sanction", "") if com else "",
        "AvisCommission": getattr(com, "AvisCommission", "") if com else "",
        "Note": getattr(com, "Note", "") if com else "",
        "Decision": getattr(com, "Decision", "") if com else "",
        "MotifDecision": getattr(com, "MotifDecision", "") if com else "",
        "Caractere": getattr(com, "Caractere", "") if com else "",
    }

    ctx = {
        "v": v,
        "agent": agent,
        "commission": commission,
        "commission_type": "",  # "Comité de relève" / "Comité de carrière" si tu veux l’indiquer
    }
    return _render_html_to_pdf(
        "vivier/pv_docx_like.html",
        ctx,
        filename=f"PV_{v.NumCommission.replace('/','_')}_{matricule}.pdf",
    )


def vivier_pv_non_traj(request, pk):
    v = get_object_or_404(Vivier.objects.using("default"), pk=pk)

    rows = []
    eligibles = compute_eligibles_agence(v.FonctionCible) if v.FonctionCible else []
    for r in eligibles:
        ag = r.get("agent")
        if not ag or bool(r.get("trajectoire")):
            continue
        com = (
            Commission.objects.using("default")
            .filter(Vivier=v, Matricule=str(ag.Matricule))
            .first()
        )
        rows.append(
            {
                "matricule": ag.Matricule,
                "nom": ag.Nom,
                "prenom": ag.Prenom,
                "fonction": getattr(ag, "FonctionLibelle", "") or "",
                "affectation_code": getattr(ag, "AffectationCode", "") or "",
                "affectation_libelle": getattr(ag, "AffectationLibelle", "") or "",
                "reseau": extract_direction_from_arbo(
                    getattr(ag, "ArborescenceAffectation", "") or ""
                ),
                "sanction": getattr(com, "Sanction", "") or "",
                "pi_n1": getattr(com, "PI_n_1", "") or "",
                "pi_n2": getattr(com, "PI_n_2", "") or "",
                "pi_n3": getattr(com, "PI_n_3", "") or "",
                "avis": getattr(com, "AvisCommission", "") or "",
                "note": getattr(com, "Note", "") or "",
                "decision": getattr(com, "Decision", "") or "",
                "motif": getattr(com, "MotifDecision", "") or "",
                "caractere": getattr(com, "Caractere", "") or "",
            }
        )

    ctx = {"v": v, "rows": rows}
    return _render_html_to_pdf(
        "vivier/pv_non_traj_pdf.html", ctx, filename=f"PV_Comite_{v.NumCommission}.pdf"
    )


MARGIN_L = 15 * mm
MARGIN_R = 15 * mm
MARGIN_T = 15 * mm
MARGIN_B = 15 * mm
PAGE_W, PAGE_H = A4


def _pdf_response(draw_pages_fn, filename="document.pdf"):
    """Crée un PDF en mémoire, laisse draw_pages_fn(canvas) dessiner les pages, renvoie HttpResponse."""
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    c.setTitle(filename)
    draw_pages_fn(c)  # <- à toi de dessiner (peut faire multi-page)
    c.save()
    buf.seek(0)
    resp = HttpResponse(buf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp


from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from .models import Vivier


@login_required
def export_agents_excel(request, pk):
    # imports locaux (aucune autre ligne de ton fichier n'est modifiée)
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # On reprend EXACTEMENT la même logique que dans vivier_update pour construire "agents"
    v = get_object_or_404(Vivier.objects.using("default"), pk=pk)

    agents = []
    if v.FonctionCible:
        try:
            raw = compute_eligibles_agence(v.FonctionCible)
        except Exception:
            raw = []

        for r in raw:
            ag = r.get("agent")
            if not ag:
                continue
            com = (
                Commission.objects.using("default")
                .filter(Vivier=v, Matricule=str(ag.Matricule))
                .first()
            )
            agents.append(
                {
                    "matricule": ag.Matricule,
                    "nom": ag.Nom,
                    "prenom": ag.Prenom,
                    "date_entree": getattr(ag, "DateEntree", None),
                    "anciennete_cpm": getattr(ag, "AncienneteAcquiseCPM", None),
                    "fonction_code": getattr(ag, "FonctionCode", "") or "",
                    "fonction_libelle": getattr(ag, "FonctionLibelle", "") or "",
                    "date_fonction": getattr(ag, "DateEffetFonction", None),
                    "anciennete_fonction": getattr(ag, "DateEffetFonction", None),
                    "aff_code": getattr(ag, "AffectationCode", "") or "",
                    "aff_lib": getattr(ag, "AffectationLibelle", "") or "",
                    "reseau": extract_direction_from_arbo(
                        getattr(ag, "ArborescenceAffectation", "") or ""
                    ),
                    "trajectoire": bool(r.get("trajectoire")),
                    "exception": bool(r.get("exception")),
                    "sanction": getattr(com, "Sanction", "") or "",
                    "pi_n1": getattr(com, "PI_n_1", "") or "",
                    "pi_n2": getattr(com, "PI_n_2", "") or "",
                    "pi_n3": getattr(com, "PI_n_3", "") or "",
                    "avis": getattr(com, "AvisCommission", "") or "",
                    "note": getattr(com, "Note", "") or "",
                    "decision": getattr(com, "Decision", "") or "",
                }
            )

    # Filtre "traj" identique à la page
    traj = (request.GET.get("traj") or "").lower()
    if traj == "oui":
        agents = [a for a in agents if a["trajectoire"]]
    elif traj == "non":
        agents = [a for a in agents if a["exception"]]

    # En-têtes alignés sur ton tableau
    headers = [
        "Matricule",
        "Nom",
        "Prénom",
        "Date d'entrée",
        "Ancienneté CPM",
        "Fonction Code",
        "Fonction Libellé",
        "Date fonction occupée",
        "Anc. fonction",
        "Affectation Code",
        "Affectation Libellé",
        "Réseau",
        "Trajectoire",
        "Sanction",
        "PI N-1",
        "PI N-2",
        "PI N-3",
        "Avis de la commission",
        "Note",
        "Décision",
    ]

    wb = Workbook()
    ws = wb.active

    # 🔒 Titre d’onglet : nettoyer caractères interdits Excel et limiter à 31
    raw_title = f"Vivier {v.NumCommission}"
    forbidden = set(r":\/?*[]")
    safe_title = "".join("-" if ch in forbidden else ch for ch in raw_title).strip()
    if not safe_title:
        safe_title = "Vivier"
    ws.title = safe_title[:31]

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # En-têtes
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.freeze_panes = "A2"

    max_len = [len(h) for h in headers]
    row_idx = 2
    for a in agents:
        row = [
            a["matricule"] or "",
            a["nom"] or "",
            a["prenom"] or "",
            a["date_entree"] or "",
            a["anciennete_cpm"] or "—",
            a["fonction_code"] or "",
            a["fonction_libelle"] or "",
            a["date_fonction"] or "",
            a["anciennete_fonction"] or "—",
            a["aff_code"] or "",
            a["aff_lib"] or "",
            a["reseau"] or "",
            ("Oui" if a["trajectoire"] else "Non"),
            a["sanction"] or "—",
            a["pi_n1"] or "—",
            a["pi_n2"] or "—",
            a["pi_n3"] or "—",
            a["avis"] or "—",
            a["note"] or "—",
            {"RETENU": "Retenu(e)", "NON_RETENU": "Non retenu(e)"}.get(
                a["decision"], "-"
            ),
        ]
        ws.append(row)

        # bordures
        for cell in ws[row_idx]:
            cell.border = border
        # centrages utiles
        for col in (1, 4, 8, 13, 19, 20):
            ws.cell(row=row_idx, column=col).alignment = center

        # largeur auto
        for i, val in enumerate(row):
            s = val.strftime("%d/%m/%Y") if hasattr(val, "strftime") else str(val)
            max_len[i] = max(max_len[i], len(s))

        row_idx += 1

    # Ajuster largeurs
    for i, width in enumerate(max_len, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(
            60, max(10, width + 2)
        )

    # Format date (colonnes 4 et 8)
    date_fmt = "DD/MM/YYYY"
    for col in (4, 8):
        for r in range(2, row_idx):
            cell = ws.cell(row=r, column=col)
            if hasattr(cell.value, "strftime"):
                cell.number_format = date_fmt

    # Réponse HTTP
    now = timezone.now().strftime("%Y%m%d_%H%M%S")
    safe_num = str(v.NumCommission).replace("/", "_")
    filename = f"Vivier_{safe_num}_Agents_{now}.xlsx"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def draw_wrapped(c, text, x, y, max_width, leading, font=BASE_FONT, size=10):
    """Dessine du texte avec retour à la ligne pour ne pas dépasser max_width. Retourne le y final."""
    if not text:
        return y
    words = str(text).split()
    line, lines = "", []
    for w in words:
        test = (line + " " + w).strip()
        if stringWidth(test, font, size) <= max_width:
            line = test
        else:
            if line:
                lines.append(line)
            line = w
    if line:
        lines.append(line)
    c.setFont(font, size)
    cur_y = y
    for ln in lines:
        c.drawString(x, cur_y, ln)
        cur_y -= leading
    return cur_y


def draw_header_footer(c, title_left="", title_right=""):
    """En-tête + pied de page uniformes."""
    # Header line
    c.setLineWidth(0.6)
    c.line(MARGIN_L, PAGE_H - MARGIN_T, PAGE_W - MARGIN_R, PAGE_H - MARGIN_T)
    c.setFont(BASE_FONT, 11)
    if title_left:
        c.drawString(MARGIN_L, PAGE_H - MARGIN_T + 4 * mm, title_left)
    if title_right:
        c.drawRightString(PAGE_W - MARGIN_R, PAGE_H - MARGIN_T + 4 * mm, title_right)
    # Footer line + page number
    c.line(MARGIN_L, MARGIN_B, PAGE_W - MARGIN_R, MARGIN_B)
    c.setFont(BASE_FONT, 9)
    c.drawRightString(PAGE_W - MARGIN_R, MARGIN_B - 4 * mm, f"Page {c.getPageNumber()}")


def new_page(c, title_left="", title_right=""):
    c.showPage()
    draw_header_footer(c, title_left, title_right)


def ensure_space(c, y, needed):
    """Retourne (y, page_changed) en créant une nouvelle page si besoin."""
    if y - needed < MARGIN_B + 10:  # marge basse + petit padding
        new_page(c, getattr(c, "_title_left", ""), getattr(c, "_title_right", ""))
        return PAGE_H - MARGIN_T - 10, True
    return y, False
