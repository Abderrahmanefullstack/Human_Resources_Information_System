from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from pandas import qcut
from .models import Fonction
from django.http import HttpResponseRedirect
from django.db import transaction
from unidecode import unidecode
from django.urls import reverse
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.utils import timezone

from django.contrib import messages
from difflib import SequenceMatcher
from django.shortcuts import redirect
from agents.views import clean_column
from fonctions.forms import FonctionForm
from fonctions.models import Fonction
from django.db.models import Q
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
import pandas as pd
import logging

logger = logging.getLogger(__name__)

# def fonctions_list(request):
# ici tu récupères les données nécessaires
# return render(request, "fonctions/fonction_list.html", {})


def fonction_detail(request, code):
    fonction = get_object_or_404(Fonction, Code=code)
    return render(request, "fonctions/fonction_detail.html", {"fonction": fonction})


@login_required
def export_excel(request):
    q = request.GET.get("q", "").strip()

    qs = Fonction.objects.all()
    if q:
        qs = qs.filter(
            Q(Code__icontains=q)
            | Q(Intitule__icontains=q)
            | Q(Intitule_Complet__icontains=q)
        )

    headers = ["Code", "Intitulé", "Intitulé Complet"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Fonctions"

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # En-têtes
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    ws.freeze_panes = "A2"

    max_len = [len(h) for h in headers]
    row_idx = 2

    for f in qs.only("Code", "Intitule", "Intitule_Complet"):
        row = [f.Code or "", f.Intitule or "", f.Intitule_Complet or ""]
        ws.append(row)

        for cell in ws[row_idx]:
            cell.border = border
        ws.cell(row=row_idx, column=1).alignment = center  # Code centré
        row_idx += 1

        # Largeurs auto
        for i, val in enumerate(row):
            max_len[i] = max(max_len[i], len(str(val)))

    # Ajuster largeurs
    for i, width in enumerate(max_len, start=1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = min(60, max(10, width + 2))

    now = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Fonctions_{now}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def import_excel(request):
    # ⭐ Toujours renvoyer une réponse même hors POST
    if request.method != "POST":
        messages.error(request, "Méthode invalide.")
        return redirect("fonctions:list")

    if not request.FILES.get("excel_file"):
        messages.error(request, "⚠️ Aucun fichier sélectionné.")
        return redirect("fonctions:list")

    excel_file = request.FILES["excel_file"]

    try:
        # Lis en texte pour préserver les codes
        df = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
        print("✅ Fichier Excel chargé :", df.shape)

        # Nettoyage des colonnes Excel
        df.columns = [clean_column(col) for col in df.columns]
        print("🧼 Colonnes Excel nettoyées :", df.columns[:10])

        # 🔧 Mappage forcé
        forced_mapping = {
            "Code": "CODE",
            "Intitule": "INTITULE",
            "Intitule_Complet": "INTITULECOMPLET",
        }
        matched_fields = {m: e for m, e in forced_mapping.items() if e in df.columns}
        print("🔗 Champs appariés :", matched_fields)

        if "Code" not in matched_fields:
            messages.error(request, "⚠️ La colonne 'Code' est obligatoire.")
            return redirect("fonctions:list")

        # Construire avant suppression
        fonctions = []
        seen_codes = set()
        code_col = matched_fields.get("Code")
        inti_col = matched_fields.get("Intitule")
        inti_cpl_col = matched_fields.get("Intitule_Complet")

        for row in df.itertuples(index=False):
            code_val = getattr(row, code_col, None)
            if code_val is None or (isinstance(code_val, float) and pd.isna(code_val)):
                continue
            code_str = str(code_val).strip()
            if not code_str or code_str in seen_codes:
                continue
            seen_codes.add(code_str)

            inti_val = getattr(row, inti_col, "") if inti_col else ""
            inti_cpl_val = getattr(row, inti_cpl_col, "") if inti_cpl_col else ""

            fonctions.append(
                Fonction(
                    Code=code_str,
                    Intitule=str(inti_val).strip() if inti_val is not None else "",
                    Intitule_Complet=(
                        str(inti_cpl_val).strip() if inti_cpl_val is not None else ""
                    ),
                )
            )

        if not fonctions:
            messages.error(
                request, "⚠️ Aucun enregistrement valide trouvé (colonne 'Code')."
            )
            return redirect("fonctions:list")

        with transaction.atomic():
            Fonction.objects.all().delete()
            print("🗑️ Données Fonction supprimées.")
            Fonction.objects.bulk_create(fonctions, batch_size=1000)

        print(f"✅ {len(fonctions)} fonctions insérées.")
        messages.success(request, f"{len(fonctions)} fonctions importées avec succès.")
        return redirect("fonctions:list")

    except Exception as e:
        logger.exception("❌ Erreur lors de l'import Excel")
        messages.error(request, f"Erreur lors de l'import : {str(e)}")
        return redirect("fonctions:list")


@login_required
def fonction_list(request):
    query = request.GET.get("q", "").strip()
    fonctions = Fonction.objects.all()
    filters = {}

    if query:
        fonctions = fonctions.filter(
            Q(Code__icontains=query)
            | Q(Intitule__icontains=query)
            | Q(Intitule_Complet__icontains=query)
        )
        filters["q"] = query

    context = {
        "fonctions": fonctions,
        "filters": filters,
    }
    return render(request, "fonctions/fonction_list.html", context)


def fonction_update(request, code):
    fonction = get_object_or_404(Fonction, Code=code)
    if request.method == "POST":
        form = FonctionForm(request.POST, instance=fonction)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Fonction modifiée avec succès.")
            return redirect("fonctions:list")
    else:
        form = FonctionForm(instance=fonction)
    return render(request, "fonctions/fonction_form.html", {"form": form, "edit": True})


def fonction_delete(request, code):
    fonction = get_object_or_404(Fonction, Code=code)
    if request.method == "POST":
        fonction.delete()
        messages.success(request, "🗑️ Fonction supprimée.")
        return redirect("fonctions:list")
    return render(
        request, "fonctions/fonction_confirm_delete.html", {"fonction": fonction}
    )


def download_template(request):
    # Création d’un DataFrame avec les colonnes de base
    df = pd.DataFrame(columns=["CODE", "INTITULE", "INTITULE_COMPLET"])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=modele_fonctions.xlsx"

    with pd.ExcelWriter(response, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Fonctions")

    return response


def get_db_columns():
    return {
        "CODE",
        "INTITULE",
        "INTITULE_COMPLET",
    }


# Colonnes utiles (en minuscule pour le nettoyage)
EXPECTED_COLUMNS = {
    "CODE",
    "INTITULE",
    "INTITULE_COMPLET",
}


def parse_boolean(value):
    if isinstance(value, str):
        value = value.strip().upper()
        if value in ["OUI", "YES", "TRUE", "1"]:
            return True
        elif value in ["NON", "NO", "FALSE", "0"]:
            return False
    elif isinstance(value, (int, float)):
        return value == 1
    return None


def clean_column(col):
    return (
        unidecode(str(col))
        .upper()
        .replace(" ", "")
        .replace("'", "")
        .replace("’", "")
        .replace("(", "")
        .replace(")", "")
        .replace("-", "")
        .replace("_", "")
    )


def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()


def purge_fonctions(request):
    deleted_count = Fonction.objects.using("default").all().delete()
    return HttpResponse(f"🗑️ {deleted_count[0]} fonctions supprimées")


def normalize_name(name):
    return re.sub(r"[^a-zA-Z0-9]", "", name).lower()
