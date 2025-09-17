from django.shortcuts import render, get_object_or_404, redirect
from .models import Agent
from entites.models import Entite
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.contrib import messages
from .forms import AgentForm
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re
from django.http import HttpResponse
from .forms import ExcelUploadForm
from django.db import connection
from tqdm import tqdm
import logging
from django.db import transaction
from unidecode import unidecode
from difflib import SequenceMatcher
from agents.models import ImportLog
from django.utils import timezone
from datetime import datetime, timedelta
from django.db.models.functions import ExtractYear
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP


def test_connection(request):
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT TOP 5 Matricule FROM Agent")
            rows = cursor.fetchall()
            return HttpResponse(str(rows))
    except Exception as e:
        return HttpResponse(f"Erreur SQL: {str(e)}")


@login_required
def agent_detail(request, pk):
    agent = get_object_or_404(Agent, pk=pk)
    return render(request, "agents/agent_detail.html", {"agent": agent})


@login_required
def agent_update(request, pk):
    agent = get_object_or_404(Agent, pk=pk)
    if request.method == "POST":
        form = AgentForm(request.POST, instance=agent)
        if form.is_valid():
            form.save()
            messages.success(request, "Agent modifié avec succès.")
            return HttpResponseRedirect(reverse("agents:list"))
    else:
        form = AgentForm(instance=agent)
    return render(request, "agents/agent_form.html", {"form": form})


logger = logging.getLogger(__name__)


def agent_delete(request, pk):
    agent = get_object_or_404(Agent, pk=pk)

    if request.method == "POST":
        try:
            agent.delete()
            logger.warning(f"Agent supprimé : {agent.Matricule} par {request.user}")
            messages.success(request, "✅ Suppression confirmée.")
            return redirect("agents:list")  # redirection vers la liste
        except Exception as e:
            logger.error(f"❌ Erreur lors de la suppression : {str(e)}")
            messages.error(request, f"Erreur : {e}")
            return redirect("agents:detail", pk=pk)  # retour vers détail si erreur

    # Méthode GET : afficher la page de confirmation
    return render(request, "agents/agent_confirm_delete.html", {"agent": agent})


@login_required
def agent_list(request):
    query = request.GET.get("q", "").strip()
    genre = request.GET.get("genre", "").upper()
    annee_retraite = request.GET.get("annee_retraite", "").strip()
    print("🔎 Année sélectionnée :", annee_retraite)

    agents = Agent.objects.all()
    filters = {}

    # 🔍 Filtre recherche (nom, matricule...)
    if query:
        agents = agents.filter(
            Q(Matricule__icontains=query)
            | Q(Nom__icontains=query)
            | Q(Prenom__icontains=query)
            | Q(Civilite__icontains=query)
            | Q(NoAffiliationCNSS__icontains=query)
            | Q(AffectationCode__icontains=query)
            | Q(AffectationLibelle__icontains=query)
            | Q(ArborescenceAffectation__icontains=query)
        )
        filters["q"] = query

    # 🎭 Filtre genre
    if genre == "H":
        agents = agents.filter(Civilite__iexact="Mr")
        filters["genre"] = "H"
    elif genre == "F":
        agents = agents.filter(Civilite__in=["Mme", "Melle"])
        filters["genre"] = "F"
    else:
        filters["genre"] = ""

    # 🎂 Filtre retraite (à appliquer en dernier)
    if annee_retraite:
        try:
            annee_retraite = int(annee_retraite)
            print("📅 Année sélectionnée :", annee_retraite)

            agents = agents.filter(DateNaissance__isnull=False)
            print("🧠 Agents avec DateNaissance connue :", agents.count())

            # Transformation en liste et filtrage Python
            agents = [
                agent
                for agent in agents
                if (agent.DateNaissance + timedelta(days=60 * 365.25)).year
                == annee_retraite
            ]
            print(
                "🎯 Agents prenant leur retraite en", annee_retraite, ":", len(agents)
            )

            filters["annee_retraite"] = annee_retraite

        except Exception as e:
            print("❌ Erreur dans le filtre retraite :", e)

    # 🆕 Récupération des derniers imports
    recent_imports = ImportLog.objects.order_by("-import_date")[:5]

    context = {
        "agents": agents,
        "filters": filters,
        "recent_imports": recent_imports,
        "years": list(range(2025, 2071)),
    }

    return render(request, "agents/agents_list.html", context)


def normalize_name(name):
    return re.sub(r"[^a-zA-Z0-9]", "", name).lower()


def download_template(request):
    try:
        # 1. Récupération dynamique des colonnes
        db_columns = sorted(get_db_columns())

        # 2. Création d’un DataFrame vide avec ces colonnes
        df = pd.DataFrame(columns=db_columns)

        # 3. Création de la réponse HTTP avec le fichier Excel
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="modele_import_agents.xlsx"'
        )
        df.to_excel(response, index=False, engine="openpyxl")

        return response

    except Exception as e:
        messages.error(request, f"Erreur lors du téléchargement du modèle : {str(e)}")
        return redirect("agents:list")

    # Avant pd.read_excel(...)

    if not excel_file.name.lower().endswith(".xlsx"):
        messages.error(
            request, "❌ Format non conforme. Merci d'importer un fichier .xlsx"
        )
        return redirect("agents:list")

    try:
        df = pd.read_excel(excel_file, engine="openpyxl")
    except ImportError:
        messages.error(
            request,
            "❌ Dépendance manquante: installe openpyxl (pip install openpyxl).",
        )
        return redirect("agents:list")

    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]

        try:
            print("DEBUG: Fichier reçu → démarrage import")

            # 1. Lecture du fichier avec gestion des erreurs
            try:
                df = pd.read_excel(excel_file, engine="openpyxl")
                print("DEBUG: Fichier Excel lu avec succès, lignes :", len(df))
            except Exception as e:
                messages.error(request, f"Format de fichier invalide : {str(e)}")
                print("ERREUR: Lecture Excel échouée :", str(e))
                return redirect("agents:list")

            # 2. Nettoyage des noms de colonnes
            original_columns = list(df.columns)
            cleaned_columns = [clean_column(col) for col in df.columns]
            df.columns = cleaned_columns

            print("DEBUG: Colonnes Excel nettoyées :", set(cleaned_columns))

            # 3. Vérification avec les colonnes attendues
            db_columns = get_db_columns()
            print("DEBUG: Colonnes attendues BDD    :", db_columns)

            excel_columns = set(df.columns)

            if not db_columns.issubset(excel_columns):
                missing = db_columns - excel_columns
                extra = excel_columns - db_columns

                error_msg = [
                    "<div class='text-start'>",
                    "<strong>Erreur de validation :</strong>",
                    "<ul class='mb-0'>",
                ]
                if missing:
                    error_msg.append(
                        f"<li>Colonnes manquantes : <code>{', '.join(sorted(missing))}</code></li>"
                    )
                if extra:
                    error_msg.append(
                        f"<li>Colonnes inattendues : <code>{', '.join(sorted(extra))}</code></li>"
                    )
                error_msg.append("</ul></div>")

                messages.error(request, "\n".join(error_msg))
                print("ERREUR: Problème de colonnes")
                return redirect("agents:import_excel")

            # 4. Écrasement des données existantes + insertion
            with transaction.atomic():
                print("DEBUG: Suppression des agents existants...")
                deleted = Agent.objects.using("default").all().delete()
                print(f"DEBUG: {deleted[0]} agents supprimés")

                agents_to_create = []
                for _, row in tqdm(df.iterrows(), total=len(df), desc="Importation"):
                    agent_data = {field: row.get(field) for field in db_columns}
                    agents_to_create.append(Agent(**agent_data))

                Agent.objects.using("default").bulk_create(agents_to_create)
                print(f"DEBUG: {len(agents_to_create)} agents insérés.")

            # 5. Message de succès
            success_msg = f"""
            <div class='alert-icon'>
                <i class='fas fa-check-circle me-2'></i>
                <strong>Importation réussie !</strong>
            </div>
            <div class='mt-2 text-start'>
                <span class='badge bg-success'>{len(df)} lignes importées</span>
                <small class='d-block mt-1'>Dernier matricule importé : {df.iloc[-1].get("MATRICULE", "Inconnu")}</small>
            </div>
            """
            messages.success(request, success_msg)
            logger.info(f"Import réussi - {len(df)} agents - Par {request.user}")

        except Exception as e:
            logger.error(f"Échec import : {str(e)}", exc_info=True)
            print("ERREUR IMPORT CRITIQUE :", str(e))
            messages.error(
                request,
                f"""
                <div class='alert-icon'>
                    <i class='fas fa-exclamation-triangle me-2'></i>
                    <strong>Erreur critique</strong>
                </div>
                <div class='mt-2'>
                    <code>{str(e)}</code>
                </div>
                """,
            )

    return redirect("agents:list")


@login_required
def export_excel(request):
    # Reprend tes filtres EXACTS comme dans agent_list
    query = request.GET.get("q", "").strip()
    genre = request.GET.get("genre", "").upper()
    annee_retraite = request.GET.get("annee_retraite", "").strip()

    # Base queryset
    qs = Agent.objects.all()

    if query:
        qs = qs.filter(
            Q(Matricule__icontains=query)
            | Q(Nom__icontains=query)
            | Q(Prenom__icontains=query)
            | Q(Civilite__icontains=query)
            | Q(NoAffiliationCNSS__icontains=query)
            | Q(AffectationCode__icontains=query)
            | Q(AffectationLibelle__icontains=query)
            | Q(ArborescenceAffectation__icontains=query)
        )

    if genre == "H":
        qs = qs.filter(Civilite__iexact="Mr")
    elif genre == "F":
        qs = qs.filter(Civilite__in=["Mme", "Melle"])

    # ⚠️ Ton filtre retraite transforme en liste Python → on refait pareil
    if annee_retraite:
        try:
            annee_retraite = int(annee_retraite)
            qs = qs.filter(DateNaissance__isnull=False)
            qs = [
                a
                for a in qs
                if (a.DateNaissance + timedelta(days=60 * 365.25)).year
                == annee_retraite
            ]
        except Exception:
            pass

    # Colonnes à exporter (calquées sur ton tableau)
    headers = [
        "Matricule",
        "Civilité",
        "Nom",
        "Prénom",
        "N° CNSS",
        "Date Naissance",
        "Date d'entrée",
        "Situation Effectif",
        "Affectation Code",
        "Affectation Libellé",
        "Arborescence Affectation",
    ]

    # Prépare le classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Agents"

    # Style en-tête
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Ecrire en-têtes
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Geler l’en-tête
    ws.freeze_panes = "A2"

    # Largeurs auto
    max_len = [len(h) for h in headers]

    # Itérer les lignes (qs peut être QuerySet ou liste après filtre retraite)
    def iter_agents(seq):
        # Optimise si QuerySet (values_only)
        if not isinstance(seq, list):
            for a in seq.only(
                "Matricule",
                "Civilite",
                "Nom",
                "Prenom",
                "NoAffiliationCNSS",
                "DateNaissance",
                "DateEntree",
                "SituationEffectifLibelle",
                "AffectationCode",
                "AffectationLibelle",
                "ArborescenceAffectation",
            ):
                yield a
        else:
            for a in seq:
                yield a

    row_idx = 2
    for a in iter_agents(qs):
        # Formattage identique à ton tableau : Matricule sur 10 chiffres
        matricule = (
            (str(a.Matricule).strip().zfill(10)) if a.Matricule is not None else ""
        )

        # Dates au format Excel (laisser des objets date si possibles)
        dn = a.DateNaissance if a.DateNaissance else ""
        de = a.DateEntree if a.DateEntree else ""

        row = [
            matricule,
            a.Civilite or "",
            a.Nom or "",
            a.Prenom or "",
            a.NoAffiliationCNSS or "",
            dn,
            de,
            a.SituationEffectifLibelle or "",
            a.AffectationCode or "",
            a.AffectationLibelle or "",
            a.ArborescenceAffectation or "",
        ]

        ws.append(row)
        # Borders + alignement pour certaines colonnes
        for cell in ws[row_idx]:
            cell.border = border
        ws.cell(row=row_idx, column=1).alignment = center  # Matricule
        ws.cell(row=row_idx, column=6).alignment = center  # Date Naissance (affichage)
        ws.cell(row=row_idx, column=7).alignment = center  # Date d'entrée
        row_idx += 1

        # Largeurs
        for i, val in enumerate(row):
            s = val.strftime("%d/%m/%Y") if hasattr(val, "strftime") else str(val)
            max_len[i] = max(max_len[i], len(s))

    # Ajuster largeurs (cap à 60)
    for i, width in enumerate(max_len, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(
            60, max(10, width + 2)
        )

    # Forcer un format date lisible (si tu veux du texte, on peut changer, mais je ne renomme rien)
    date_fmt = "DD/MM/YYYY"
    for col in (6, 7):
        for r in range(2, row_idx):
            cell = ws.cell(row=r, column=col)
            if hasattr(cell.value, "strftime"):
                cell.number_format = date_fmt

    # Retour HTTP
    now = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Agents_{now}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def get_db_columns():
    return {
        "MATRICULE",
        "CIVILITE",
        "NOM",
        "PRENOM",
        "NOAFFILIATIONCNSS",
        "DATEENTREE",
        "SITUATIONEFFECTIFLIBELLE",
        "AFFECTATIONCODE",
        "AFFECTATIONLIBELLE",
        "ARBORESCENCEAFFECTATION",
    }


logger = logging.getLogger(__name__)

# Colonnes utiles (en minuscule pour le nettoyage)
EXPECTED_COLUMNS = {
    "matricule",
    "civilite",
    "nom",
    "prenom",
    "noaffiliationcnss",
    "dateentree",
    "situationeffectiflibelle",
    "affectationcode",
    "affectationlibelle",
    "arborescenceaffectation",
}


def parse_boolean(value):
    if isinstance(value, str):
        value = value.strip().upper()
        if value in ["OUI", "YES", "TRUE", "1"]:
            return True
        elif value in ["NON", "NO", "FALSE", "0"]:
            return False
    elif isinstance(value, (int, float)):
        return value == 1
    return None


logger = logging.getLogger(__name__)


def clean_column(col):
    return (
        unidecode(str(col))
        .upper()
        .replace(" ", "")
        .replace("'", "")
        .replace("’", "")
        .replace("(", "")
        .replace(")", "")
        .replace("-", "")
        .replace("_", "")
    )


def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()


def import_excel(request):
    if request.method != "POST" or not request.FILES.get("excel_file"):
        messages.warning(request, "Aucun fichier sélectionné.")
        return redirect("agents:list")

    excel_file = request.FILES["excel_file"]

    # --- petit helper local pour normaliser proprement les valeurs texte ---
    def _norm(v, allow_empty=False):
        if v in (None, ""):
            return "" if allow_empty else None
        s = str(v).strip()
        if not s and not allow_empty:
            return None
        return s

    try:
        df = pd.read_excel(excel_file, engine="openpyxl")
        print("Fichier Excel chargé :", df.shape)

        # 1) Nettoyage colonnes
        df.columns = [clean_column(col) for col in df.columns]
        print("Colonnes Excel nettoyées :", df.columns[:10])

        # 2) Appariement colonnes -> champs modèle
        model_fields = {f.name: clean_column(f.name) for f in Agent._meta.fields}
        matched_fields = {}

        forced_mapping = {
            # (ton mapping forcé corrigé/fiable)
            "Matricule": "MATRICULE",
            "Civilite": "CIVILITE",
            "Nom": "NOM",
            "Prenom": "PRENOM",
            "NoAffiliationCNSS": "NOAFFILIATIONCNSS",
            "DateEntree": "DATEENTREE",
            "SituationEffectifLibelle": "SITUATIONEFFECTIFLIBELLE",
            "AffectationCode": "AFFECTATIONCODE",
            "AffectationLibelle": "AFFECTATIONLIBELLE",
            "ArborescenceAffectation": "ARBORESCENCEAFFECTATION",
            "NationaliteCode": "NATIONALITECODE",
            "NationaliteLibelle": "NATIONALITELIBELLE",
            "DateNaissance": "DATENAISSANCE",
            "DateNaissanceCNSS": "DATENAISSANCECNSS",
            "LieuNaissance": "LIEUNAISSANCE",
            "SituationFamilleCode": "SITUATIONDEFAMILLECODE",
            "SituationFamilleLibelle": "SITUATIONDEFAMILLELIBELLE",
            "DateSituationFamille": "DATESITUATIONDEFAMILLE",
            "DroitPaie": "DROITLAPAIEOUINON",
            "DateDeces": "DATEDEDCS",
            "DateCertificatDeces": "DATECERTIFICATDEDCS",
            "ChefDeFamille": "CHEFDEFAMILLEOUINON",
            "ConjointCPM": "CONJOINTCPMOUINON",
            "NombreEnfants": "NOMBREDENFANTS",
            "NombreEnfantsCharge": "NOMBREDENFANTSCHARGE",
            "NombreDeductions": "NOMBREDEDDUCTIONS",
            "NoCIN": "NOCIN",
            "DateExpirationCIN": "DATEDEXPIRATIONCIN",
            "AutoriteDelivranceCIN": "AUTORITDLIVRANCECIN",
            "NoPassePort": "NOPASSEPORT",
            "DateDelivrancePassePort": "DATEDLIVRANCEPASSEPORT",
            "DateExpirationPassePort": "DATEXPIRATIONPASSEPORT",
            "AutoriteDelivrancePassePort": "AUTORITDLIVRANCEPASSEPORT",
            "RegimeCNSS": "RGIMECNSS",
            "DateAffiliationCNSS": "DATEAFFILIATIONCNSS",
            "RegimeCIMR": "RGIMECIMR",
            "NoAffiliationCIMR": "NOAFFILIATIONCIMR",
            "DateAffiliationCIMR": "DATEAFFILIATIONCIMR",
            "RegimeMUTUELLE": "RGIMEMUTUELLE",
            "NoAffiliationMUTUELLE": "NOAFFILIATIONMUTUELLE",
            "TauxSurcotisationMutuelle": "TAUXDECOTISATIONLAMUTUELLE",
            "RegimeRCPCPM": "REGIMERCPCPM",
            "NoAffiliationRCPCPM": "NOAFFILIATIONRCPCPM",
            "MontantRETSA": "MONTANTRETSA",
            "StatutOrigine": "STATUTORIGINE",
            "TypeContratLibelle": "TYPEDECONTRATLIBELLE",
            "TypeContratCode": "TYPEDECONTRATCODE",
            "MotifEmbaucheLibelle": "MOTIFDEMBAUCHELIBELLE",
            "DateDebutAdministration": "DATEDEBUTADMINISTRATION",
            "AncienneteSecteurBancaire": "ANCIENNETESECTEURBANCAIRE",
            "NombreAnneesInterruption": "NOMBREANEESINTERUPTION",
            "AncienneteAutresSecteurs": "ANCIENNTAUTRESSECTEURS",
            "AncienneteAcquiseCPM": "ANCIENNTACQUISECPM",
            "DateDebutAnciennete": "DATEDBUTANCIENNET",
            "DateFinAdministration": "DATEFINADMINISTRATION",
            "DateFinEffectif": "DATEFINEFFECTIF",
            "MotifDepartCode": "MOTIFDEDPARTCODE",
            "MotifDepartLibelle": "MOTIFDEDPARTLIBELLE",
            "SituationEffectifCode": "SITUATIONAEFFECTIFCODE",
            "DateDebutSituationEffectif": "DATEDBUTSITUATIONALEFFECTIF",
            "DateFinSituationEffectif": "DATEFINSITUATIONALEFFECTIF",
            "Indice": "INDICE",
            "DateEffetIndice": "DATEEFFETINDICE",
            "ConventionCollectifCode": "CONVENTIONCOLLECTIFCODE",
            "ConventionCollectifLibelle": "CONVENTIONCOLLECTIFLIBELLE",
            "ClasseCode": "CLASSECODE",
            "ClasseLibelle": "CLASSELIBELLE",
            "EchelonCode": "ECHELONCODE",
            "EchelonLibelle": "ECHELONLIBELLE",
            "CategorieCode": "CATGORIECODE",
            "CategorieLibelle": "CATGORIELIBELLE",
            "GradeHierarchieCode": "GRADEHIRARCHIECODE",
            "GradeHierarchieLibelle": "GRADEHIRARCHIELIBELLE",
            "DateEffetGrade": "DATEEFFETGRADE",
            "FonctionCode": "FONCTIONCODE",
            "FonctionLibelle": "FONCTIONLIBELLE",
            "DateEffetFonction": "DATEEFFETFONCTION",
            "EmploiCode": "EMPLOICODE",
            "EmploiLibelle": "EMPLOILIBELLE",
            "LigneManagerialCode": "LIGNEMANAGRIALCODE",
            "LigneManagerialLibelle": "LIGNEMANAGRIALLIBELLE",
            "DateEffetLigneManagerial": "DATEEFFETLIGNEMANAGRIAL",
            "TypeActivite": "TYPEACTIVIT",
            "FonctionParInterimCode": "FONCTIONPARINTRIMCODE",
            "FonctionParInterimLibelle": "FONCTIONPARINTRIMLIBELLE",
            "DateDebutFonctionParInterim": "DATEDBUTFONCTIONPARINTRIM",
            "DateFinFonctionParInterim": "DATEFINFONCTIONPARINTRIM",
            "NoteN": "NOTEN",
            "NoteN_1": "NOTEN1",
            "NoteN_2": "NOTEN2",
            "NotePrimeInteressement": "NOTEPRIMEDINTERESSEMENT",
            "NotePrimeBilan": "NOTEPRIMEDEBILAN",
            "NotePrimeProductivite": "NOTEPRIMEDEPRODUCTIVITE",
            "IndiceN_1": "INDICEN1",
            "IndiceN_2": "INDICEN2",
            "IndiceN_3": "INDICEN3",
            "TauxPerformanceN_1": "TAUXPERFORMANCEN1",
            "TauxPerformanceN_2": "TAUXPERFORMANCEN2",
            "TauxPerformanceN_3": "TAUXPERFORMANCEN3",
            "NombrePointsAugmentationPromotionnelle": "NOMBREDEPOINTSAUGMENTATIONPROMOTIONNELLE",
            "MoisAvancement": "MOISDAVANCEMENT",
            "SalaireBase": "SALAIREDEBASE",
            "NombrePointsBonification": "NOMBREDEPOINTSBONIFICATION",
            "DateEffetBonification": "DATEEFFETBONIFICATION",
            "PrimeTransport": "PRIMEDETRANSPORT",
            "PrimeLogement": "PRIMEDELOGEMENT",
            "FraisRepresentation": "FRAISDEREPRESENTATION",
            "AideLogement": "AIDEAULOGEMENT",
            "PrimeSpeciale": "PRIMESPCIALE",
            "PrimeEmploi": "PRIMEDEMPLOI",
            "IndemniteLoyer": "INDEMNITDELOYER",
            "IndemniteExpat": "INDEMNITEXPATRIATION",
            "PrimeProvenceSahariale": "PRIMEDEPROVENCESAHARIAL",
            "MontantAugmentation": "MONTANTAUGMENTATION",
            "DroitPrimePanier": "DROITPRIMEDEPANIER",
            "BrutAnnuelTheorique": "BRUTANNUELTHORIQUE",
            "DernierBrutMensuel": "DERNIERBRUTMENSUEL",
            "NetAnnuelTheorique": "NETANNUELTHORIQUE",
            "DernierNetMensuel": "DERNIERNETMENSUEL",
            "BanqueAffectationCode": "BANQUEAFFECTATIONCODE",
            "BanqueAffectationLibelle": "BANQUEAFFECTATIONLIBELLE",
            "CentreResponsabilite": "CENTREDERESPONSABILIT",
            "CentreResponsabiliteOrigine": "CENTREDERESPONSABILITDORIGINE",
            "DateDebutAffectation": "DATEDBUTAFFECTATION",
            "DateFinAffectation": "DATEFINAFFECTATION",
            "ClassificationAffectation": "CLASSIFICATIONAFFECTATION",
            "MatriculeSuperieurHierarchiqueN1": "MATRICULESUPRIEURHIERARCHIQUEN1",
            "NomPrenomSuperieurHierarchiqueN1": "NOMPRENOMSUPRIEURHIERARCHIQUEN1",
            "MatriculeSuperieurHierarchiqueN2": "MATRICULESUPRIEURHIERARCHIQUEN2",
            "NomPrenomSuperieurHierarchiqueN2": "NOMPRENOMSUPRIEURHIERARCHIQUEN2",
            "LocalisationSite": "LOCALISATIONSITE",
            "LocalisationVille": "LOCALISATIONVILLE",
            "LocalisationEtage": "LOCALISATIONETAGE",
            "LocalisationBureau": "LOCALISATIONBUREAU",
        }

        # appariement forcé + similarité
        for model_field, excel_col in forced_mapping.items():
            if excel_col in df.columns:
                matched_fields[model_field] = excel_col
            else:
                best_match = None
                best_score = 0
                for col in df.columns:
                    score = similar(excel_col, col)
                    if score > best_score:
                        best_score = score
                        best_match = col
                if best_score >= 0.8 and best_match:
                    matched_fields[model_field] = best_match

        print("Champs appariés :", matched_fields)

        if "Matricule" not in matched_fields:
            messages.error(request, "La colonne 'Matricule' est obligatoire.")
            return redirect("agents:list")

        # 3) Construire les objets Agent et collecter les Entités
        agents = []
        entites_dict = {}  # AffectationCode normalisé -> instance Entite

        for _, row in df.iterrows():
            agent_data = {}
            for model_field, excel_col in matched_fields.items():
                try:
                    value = row[excel_col]
                    if pd.isna(value):
                        value = None
                    elif isinstance(value, str):
                        s = value.strip().upper()
                        if s in ("OUI", "NON"):
                            value = s == "OUI"

                    if model_field in DATE_FIELDS:
                        value = to_date(value)
                    elif model_field in DECIMAL_FIELDS_2:
                        value = to_decimal(value, decimals=2)
                    elif model_field in INTEGER_FIELDS:
                        value = to_int(value)
                    elif isinstance(value, float) and value.is_integer():
                        value = int(value)

                    agent_data[model_field] = value
                except Exception:
                    agent_data[model_field] = None

            agents.append(Agent(**agent_data))

            # --- collecte Entité (normalisée) ---
            raw_code = agent_data.get("AffectationCode")
            raw_lib = agent_data.get("AffectationLibelle")
            raw_arbo = agent_data.get("ArborescenceAffectation")

            code = _norm(raw_code)  # None si vide, sinon str(trim)
            lib = _norm(raw_lib, allow_empty=True) or ""
            arbo = _norm(raw_arbo, allow_empty=True) or ""

            if code and code not in entites_dict:
                entites_dict[code] = Entite(
                    AffectationCode=code,
                    AffectationLibelle=lib,
                    ArborescenceAffectation=arbo,
                )

        # 4) SUPPRESSION Agents (bloc dédié)
        with transaction.atomic(using="default"):
            Agent.objects.using("default").all().delete()
        print("Données Agent supprimées.")

        # 5) INSERT MASSIF Agents (avec fallback ligne-à-ligne)
        try:
            with transaction.atomic(using="default"):
                Agent.objects.using("default").bulk_create(agents, batch_size=500)
            print(f"Bulk create OK ({len(agents)} lignes).")
        except Exception:
            print("Bulk create a échoué → DIAGNOSTIC par sauvegarde ligne par ligne…")
            ok = 0
            ko = 0
            for idx, ag in enumerate(agents, start=1):
                try:
                    with transaction.atomic(using="default"):
                        ag.save(using="default")
                    ok += 1
                except Exception as row_err:
                    ko += 1
                    print(
                        f"Ligne KO #{idx} - Matricule={getattr(ag, 'Matricule', None)}"
                    )
                    for f in [
                        "SalaireBase",
                        "PrimeTransport",
                        "PrimeLogement",
                        "FraisRepresentation",
                        "AideLogement",
                        "PrimeSpeciale",
                        "PrimeEmploi",
                        "IndemniteLoyer",
                        "IndemniteExpat",
                        "PrimeProvenceSahariale",
                        "MontantAugmentation",
                        "BrutAnnuelTheorique",
                        "DernierBrutMensuel",
                        "NetAnnuelTheorique",
                        "DernierNetMensuel",
                        "NombrePointsBonification",
                    ]:
                        print(f"   {f} = {getattr(ag, f, None)!r}")
                    print(f"   Exception: {repr(row_err)}")
            print(f"Import partiel : OK={ok}, KO={ko}")
            if ko > 0:
                messages.warning(
                    request,
                    f"Import terminé avec des erreurs : {ok} lignes insérées, {ko} en échec. Consulte le terminal pour les détails.",
                )

        # 6) UPSERT des ENTITÉS (création + mise à jour douce)
        try:
            with transaction.atomic(using="default"):
                # normaliser encore une fois (sécurité)
                norm_entites = {}
                for code, ent in entites_dict.items():
                    ncode = _norm(code)
                    if not ncode:
                        continue
                    ent.AffectationCode = ncode
                    ent.AffectationLibelle = (
                        _norm(ent.AffectationLibelle, allow_empty=True) or ""
                    )
                    ent.ArborescenceAffectation = (
                        _norm(ent.ArborescenceAffectation, allow_empty=True) or ""
                    )
                    norm_entites[ncode] = ent

                codes = list(norm_entites.keys())

                # existants
                existing_codes = set(
                    Entite.objects.using("default")
                    .filter(AffectationCode__in=codes)
                    .values_list("AffectationCode", flat=True)
                )

                # à créer
                to_create = [norm_entites[c] for c in codes if c not in existing_codes]
                if to_create:
                    Entite.objects.using("default").bulk_create(
                        to_create, batch_size=500
                    )

                # à mettre à jour (libellé/arbo si on a des valeurs non vides)
                to_update = []
                if existing_codes:
                    existing_objs = {
                        e.AffectationCode: e
                        for e in Entite.objects.using("default").filter(
                            AffectationCode__in=existing_codes
                        )
                    }
                    for c in existing_codes:
                        incoming = norm_entites.get(c)
                        if not incoming:
                            continue
                        obj = existing_objs[c]
                        if incoming.AffectationLibelle:
                            obj.AffectationLibelle = incoming.AffectationLibelle
                        if incoming.ArborescenceAffectation:
                            obj.ArborescenceAffectation = (
                                incoming.ArborescenceAffectation
                            )
                        to_update.append(obj)

                    if to_update:
                        Entite.objects.using("default").bulk_update(
                            to_update,
                            ["AffectationLibelle", "ArborescenceAffectation"],
                            batch_size=500,
                        )

            print(
                f"Entités: créées={len(to_create)} | mises à jour={len(to_update)} | total collectées={len(entites_dict)}"
            )
        except Exception as e:
            print("Erreur lors de l’insertion/mise à jour des entités :", repr(e))

        # 7) Log d’import
        ImportLog.objects.create(filename=excel_file.name, import_date=timezone.now())

        # 8) Messages
        msg = f"{len(agents)} agents importés avec succès."
        try:
            msg += f" Entités: +{len(to_create)} créées, {len(to_update)} mises à jour."
        except Exception:
            pass
        messages.success(request, msg)
        return redirect("agents:list")

    except Exception as e:
        # (pas d'émoji ici pour éviter UnicodeEncodeError en console Windows)
        logger.exception("Erreur lors de l'import Excel")
        messages.error(request, f"Erreur lors de l'import : {str(e)}")
        return redirect("agents:list")


def _upsert_entites(entites_dict):
    """
    entites_dict: {code: Entite(AffectationCode=..., AffectationLibelle=..., ArborescenceAffectation=...)}
    Retourne (nb_creees, nb_mises_a_jour)
    """
    if not entites_dict:
        return (0, 0)

    codes = list(entites_dict.keys())
    existing = {
        e.AffectationCode: e
        for e in Entite.objects.using("default").filter(AffectationCode__in=codes)
    }

    to_create, to_update = [], []
    for code, draft in entites_dict.items():
        if code in existing:
            obj = existing[code]
            changed = False
            lib = draft.AffectationLibelle or ""
            arbo = draft.ArborescenceAffectation or ""
            if (obj.AffectationLibelle or "") != lib:
                obj.AffectationLibelle = lib
                changed = True
            if (obj.ArborescenceAffectation or "") != arbo:
                obj.ArborescenceAffectation = arbo
                changed = True
            if changed:
                to_update.append(obj)
        else:
            to_create.append(draft)

    created_cnt = updated_cnt = 0
    with transaction.atomic(using="default"):
        if to_create:
            Entite.objects.using("default").bulk_create(to_create, batch_size=500)
            created_cnt = len(to_create)
        if to_update:
            Entite.objects.using("default").bulk_update(
                to_update,
                ["AffectationLibelle", "ArborescenceAffectation"],
                batch_size=500,
            )
            updated_cnt = len(to_update)

    return (created_cnt, updated_cnt)


NON_DIGIT_RE = re.compile(r"[^\d,.\-\s]")


def to_decimal(val, decimals=2, allow_none=True):
    if val is None or (isinstance(val, float) and str(val) == "nan"):
        return None if allow_none else Decimal("0")

    s = str(val).strip()
    if not s:
        return None if allow_none else Decimal("0")

    # retirer devises, lettres, etc.
    s = NON_DIGIT_RE.sub("", s)
    # normaliser espaces
    s = s.replace("\xa0", " ").replace(" ", "")

    # gestion FR "12.345,67" -> "12345.67"
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        if "," in s:
            s = s.replace(",", ".")

    try:
        d = Decimal(s)
    except InvalidOperation:
        return None if allow_none else Decimal("0")

    q = Decimal("1") / (Decimal(10) ** decimals)
    return d.quantize(q, rounding=ROUND_HALF_UP)


def to_int(val, allow_none=True):
    if val is None or (isinstance(val, float) and str(val) == "nan"):
        return None if allow_none else 0
    s = str(val).strip()
    if not s:
        return None if allow_none else 0
    s = NON_DIGIT_RE.sub("", s).replace("\xa0", " ").replace(" ", "")
    s = s.replace(",", "")
    try:
        return int(Decimal(s))
    except Exception:
        return None if allow_none else 0


def to_date(val):
    if val in (None, "", float("nan")):
        return None
    if hasattr(val, "to_pydatetime"):
        try:
            return val.to_pydatetime().date()
        except Exception:
            pass
    if isinstance(val, datetime):
        return val.date()
    try:
        return datetime.fromisoformat(str(val)).date()
    except Exception:
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(str(val), fmt).date()
            except Exception:
                continue
    return None


# ---------------------------------
# Sets de champs par type (PLACEMENT CORRECT)
# ---------------------------------
DECIMAL_FIELDS_2 = {
    "PrimeTransport",
    "PrimeLogement",
    "FraisRepresentation",
    "AideLogement",
    "PrimeSpeciale",
    "PrimeEmploi",
    "IndemniteLoyer",
    "IndemniteExpat",
    "PrimeProvenceSahariale",
    "MontantAugmentation",
    "BrutAnnuelTheorique",
    "DernierBrutMensuel",
    "NetAnnuelTheorique",
    "DernierNetMensuel",
    "SalaireBase",
}

INTEGER_FIELDS = {
    "NombrePointsBonification",
}

DATE_FIELDS = {
    "DateEffetBonification",
}


def purge_agents(request):
    deleted_count = Agent.objects.using("default").all().delete()
    return HttpResponse(f"🗑️ {deleted_count[0]} agents supprimés")
